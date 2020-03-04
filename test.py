from tqdm import tqdm
from time import sleep
from nornir import InitNornir
from nornir.plugins.inventory.simple import SimpleInventory
from nornir.plugins.connections.netmiko import Netmiko
from nornir.core.deserializer.inventory import Inventory
import yaml
import pandas as pd
from openpyxl import load_workbook
import openpyxl,os
import threading

# nr = InitNornir()
# import ipdb; ipdb.set_trace()
# alist = list(range(1,10000))
# bar = tqdm(alist)
# for letter in bar:
#     bar.set_description(f"Now get {letter}")
# a=[{'ipaddress': '22222', 'macaddress': '0800-2700-b853', 'vlan': '--', 'interface': 'GE0/1', 'aging': '18', 'type': 'D'}, {'ipaddress': '192.168.56.102', 'macaddress': '8acb-e76c-0302', 'vlan': '--', 'interface': 'GE0/1', 'aging': '1', 'type': 'D'}]
# b=[{'ipaddress': '192.168.56.101', 'macaddress': '4c1c-503a-0106', 'vlan': '1', 'interface': 'GE1/0/2', 'aging': '20', 'type': 'D'}, {'ipaddress': '192.168.56.1', 'macaddress': '0800-2700-b853', 'vlan': '1', 'interface': 'GE1/0/1', 'aging': '1', 'type': 'D'}]


# if not os.path.exists(f"dis arp.xlsx"):
#     wb = openpyxl.Workbook()
#     wb.save(f"dis arp.xlsx")
# with pd.ExcelWriter(f"dis arp.xlsx") as writer:
#     aa = pd.DataFrame(a)
#     bb = pd.DataFrame(b)
#     writer.book = load_workbook(f"dis arp.xlsx")
#     aa.to_excel(writer,sheet_name="192.168.56.1",index=False)
#     bb.to_excel(writer,sheet_name="192.168.56.2",index=False)
    
    # writer.save()
filename = "test.csv"
if ".csv" in filename:
    print(True)